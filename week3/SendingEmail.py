from openpyxl import load_workbook
import os
from NaverNewsCrawler import NaverNewsCrawler
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
import re
# 사용자에게 뉴스 키워드 검색어를 입력받는다.
user_input = input('검색어를 입력하세요:')
crawler = NaverNewsCrawler(user_input)

# 수집한 데이터를 저장할 엑셀 파일명을 input을 이용해 입력받아 crawler.get_news 부분에 넣는다
excel_filename = input('엑셀 파일명을 입력하세요:')
crawler.get_news(excel_filename + ".xlsx")


# gmail 발송 기능에 필요한 계정 정보를 입력
SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 465
SMTP_USER = 'fastcampus@gmail.com'
SMTP_PASSWORD = ''  # 보안상 쓰지 않았습니다:)


# 메일 발송에 필요한 send_mail 함수
def send_mail(name, addr, subject, contents, attachment=None):
    if not re.match('(^[a-zA-Z0-9_.-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)', addr):
        print('Wrong email')
        return

    msg = MIMEMultipart('alternative')
    if attachment:
        msg = MIMEMultipart('mixed')

    msg['From'] = SMTP_USER
    msg['To'] = addr
    msg['Subject'] = name + '님, ' + subject

    text = MIMEText(contents, _charset='utf-8')
    msg.attach(text)

    if attachment:
        from email.mime.base import MIMEBase
        from email import encoders

        file_data = MIMEBase('application', 'octect-stream')
        file_data.set_payload(open(attachment, 'rb').read())
        encoders.encode_base64(file_data)

        filename = os.path.basename(attachment)
        file_data.add_header('Content-Disposition',
                             'attachment; filename="' + filename + '"')
        msg.attach(file_data)

    smtp = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT)
    smtp.login(SMTP_USER, SMTP_PASSWORD)
    smtp.sendmail(SMTP_USER, addr, msg.as_string())
    smtp.close()


# email_list.xlsx 파일의 이름과 메일정보 가져오기
wb = load_workbook('email_list.xlsx', read_only=True)
data = wb.active

# enumerate 함수로 email_list의 데이터들을 3열 까지만 행단위로 가져와서 인덱스 번호와 컬렉션의 원소를 tuple형태로 반환
for idx, row in enumerate(data.iter_rows(max_col=3)):
    if idx == 0:  # 거기서 for문으로 첫번째 행은 머릿말이니 무시하고 넘어가게 조건문을 걸고
        continue
    receiver_name = row[1].value  # 행마다의 두번째 열에있는 이름 값을 변수에 넣기
    receiver_email = row[2].value  # 행마다의 세번째 열에있는 이메일 주소 값을 변수에 넣기
    subject = "뉴스 결과 보고 드립니다."  # 제목에 해당하는 변수
    contents = receiver_name + "님, " + user_input+"에 대한 결과입니다."  # 이메일 내용에 해당하는 변수
    attachment = excel_filename + ".xlsx"  # 만들어질 파일이름 변수

    send_mail(receiver_name, receiver_email, subject, contents, attachment)
